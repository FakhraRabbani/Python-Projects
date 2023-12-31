import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)
# print(cell.value)  # transaction_id
# print(sheet.max_row)  # 4

for row in range(2, sheet.max_row + 1):
    # print(row)  # 2, 3, 4
    cell = sheet.cell(row, 3)
    # print(cell.value)  # 5.95  6.95  7.95
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()

chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')